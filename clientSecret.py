import datetime
import sys
import time
import logging
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.common import NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import pyotp
from datetime import datetime
import os
from dotenv import load_dotenv

# Charger les variables d'environnement
load_dotenv()

# Constantes
EXCEL_FILE_PATH = os.getenv('EXCEL_FILE_PATH')
REPORT_PATH_TEMPLATE = os.path.join(os.getenv('REPORT_PATH'), 'report_{}.xlsx')
AMAZON_EMAIL = os.getenv('AMAZON_EMAIL')
AMAZON_PASSWORD = os.getenv('AMAZON_PASSWORD')
AMAZON_OTP_SECRET = os.getenv('AMAZON_OTP_SECRET')

# Configuration du logging
logging.basicConfig(level=logging.INFO)
INFO = logging.info
ERROR = logging.error
WARNING = logging.WARNING

def read_accounts_from_excel(file_path):
    """Lire les comptes et IDs depuis le fichier Excel."""
    try:
        df = pd.read_excel(file_path, usecols='A,C', names=['AccountName', 'ClientID'])
        accounts = df.to_dict('records')  # Convertir en liste de dictionnaires
        logging.info(f"Loaded {len(accounts)} accounts from Excel.")
        return accounts
    except Exception as e:
        logging.error(f"Error reading Excel file: {e}")
        return []

def setup_driver():
    """Initialise le WebDriver Chrome."""
    chrome_options = ChromeOptions()
    try:
        logging.info("Initializing WebDriver with Driver Manager...")
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
        return driver
    except WebDriverException as e:
        logging.error(f"WebDriver error during initialization: {e}")
        sys.exit(1)

def login_to_amazon(driver):
    """Connectez-vous à Amazon Vendor Central."""
    try:
        logging.info("Logging into Amazon...")
        driver.delete_all_cookies()
        driver.get('https://vendorcentral.amazon.fr/')
        driver.maximize_window()

        # Saisir le nom d'utilisateur
        username = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'ap_email')))
        username.clear()
        username.send_keys(AMAZON_EMAIL)
        driver.find_element(By.ID, 'continue').click()

        # Saisir le mot de passe
        password = driver.find_element(By.ID, 'ap_password')
        password.send_keys(AMAZON_PASSWORD)
        driver.find_element(By.ID, 'signInSubmit').click()

        # Saisir le code OTP
        otp_code = pyotp.TOTP(AMAZON_OTP_SECRET).now()
        code = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'auth-mfa-otpcode')))
        code.send_keys(otp_code)
        driver.find_element(By.ID, 'auth-signin-button').click()

        logging.info("Logged into Amazon successfully.")
    except Exception as e:
        logging.error(f"Unexpected error during login: {e}")
        driver.quit()
        sys.exit(1)

def select_client_account(driver, account_name):
    try:
        logging.info(f"Selecting account: {account_name}")
        driver.get('https://vendorcentral.amazon.fr/account-switcher/regional/vendorGroup')
        account_list = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '.full-page-account-switcher-accounts'))
        )
        time.sleep(3)
        accounts = account_list.find_elements(By.CLASS_NAME, 'full-page-account-switcher-account')
        time.sleep(2)
        logging.info(f"Found {len(accounts)} accounts in the list.")

        account_found = False

        for account in accounts:
            account_label = account.find_element(By.CLASS_NAME, 'full-page-account-switcher-account-label')
            if account_label.text.strip() == account_name.strip():
                account.find_element(By.CLASS_NAME, 'full-page-account-switcher-account-details').click()
                logging.info(f"Account {account_name} selected.")
                account_found = True
                time.sleep(2)
                break

        if not account_found:
            WARNING(f"Account {account_name} not found in the list. Moving to the next client.")
            return

        try:
            submit_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//button[contains(@class, "kat-button--primary") and contains(@class, "kat-button--base")]'))
            )
            submit_button.click()
            INFO("Submit button clicked.")
        except Exception as e:
            logging.warning(f"Submit button not found or failed to click: {e}")

        INFO(f"Account {account_name} selected and submitted successfully.")
    except Exception as e:
        ERROR(f"Error selecting account {account_name}: {e}")

def check_developer_profile_alert(driver):
    """Vérifiez si une alerte de profil développeur est présente sur la page."""
    try:
        alert_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'You need to complete your Developer Profile')]"))
        )
        return True if alert_element else False
    except (NoSuchElementException, TimeoutException):
        return False

def close_modal_if_open(driver):
    try:
        # Vérifiez si un modal est ouvert
        modal = driver.find_elements(By.CSS_SELECTOR, "kat-modal[role='dialog'][aria-modal='true']")
        if modal:
            close_button = modal[0].find_element(By.CSS_SELECTOR, "button.close")
            driver.execute_script("arguments[0].click();", close_button)
            logging.info("Modal fermé avant de continuer.")
            time.sleep(2)  # Attendre que le modal se ferme avant de cliquer à nouveau
    except Exception as e:
        logging.warning(f"Erreur lors de la fermeture du modal: {e}")

def find_application_by_client_id(driver, client_id):
    try:
        driver.get('https://vendorcentral.amazon.fr/sellingpartner/developerconsole?ref_=vc_xx_subNav')
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'applicationTableBody')))

        for row_index in range(1, 5):  # Exécute sur les quatre premières lignes (ajuster si nécessaire)
            try:
                close_modal_if_open(driver)
                view_button_xpath = f"/html/body/div[1]/div[2]/div/div/div/div/div/kat-tabs/kat-tab[1]/div[4]/kat-table[2]/kat-table-body/kat-table-row[{row_index}]/kat-table-cell[3]/div/kat-link"
                view_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, view_button_xpath)))
                view_button.click()
                logging.info(f"Clicked View button for application row {row_index}")

                # Attendre que le modal soit visible
                WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-modal[role='dialog'][aria-modal='true']"))
                )

                # Accéder au Shadow DOM pour le modal
                shadow_host = driver.find_element(By.CSS_SELECTOR, "kat-modal[role='dialog'][aria-modal='true']")
                shadow_root = driver.execute_script("return arguments[0].shadowRoot", shadow_host)

                # Localiser et extraire l'ID client depuis le modal
                client_id_element = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((By.ID, "clientIdInput"))
                )
                found_client_id = client_id_element.get_attribute("value")
                logging.info(f"ClientID trouvé: {found_client_id}")

                if found_client_id == client_id:
                    logging.info(f"Le ClientID correspond pour le compte: {client_id}")

                    # Appeler la fonction pour cliquer sur la flèche et afficher la clé secrète
                    click_arrow(driver)

                    return True
                else:
                    close_button = shadow_root.find_element(By.CSS_SELECTOR, "button.close")
                    driver.execute_script("arguments[0].click();", close_button)
                    logging.info("Modal closed.")
                    time.sleep(2)
            except Exception as e:
                logging.error(f"Erreur lors de la vérification du ClientID pour l'application à la ligne {row_index}: {e}")
                continue

        logging.warning(f"Aucun profil d'application trouvé pour le ClientID: {client_id}")
        return False
    except Exception as e:
        logging.error(f"Erreur lors de la recherche du ClientID: {e}")
        return False

def click_view_button(driver, account_name, report_data):
    try:
        driver.get('https://vendorcentral.amazon.fr/sellingpartner/developerconsole?ref_=vc_xx_subNav')
        time.sleep(2)
        if check_developer_profile_alert(driver):
            report_data.append([account_name, "N/A", "Client needs to complete their Developer Profile", "N/A", "N/A", "N/A", "N/A"])
            return

        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'applicationTableBody')))

        logging.info("Attempting to click the view link...")
        view_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[2]/div/div/div/div/div/kat-tabs/kat-tab[1]/div[4]/kat-table[2]/kat-table-body/kat-table-row/kat-table-cell[3]/div/kat-link'))
        )
        view_link.click()
        logging.info("View link clicked.")

    except Exception as e:
        logging.error(f"An error occurred while clicking view link: {e}")

def click_arrow(driver):
    """Click on the arrow to display the secret key."""
    try:
        modal_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "kat-modal[role='dialog'][aria-modal='true'][visible='true']"))
        )
        INFO("Modal visible")

        shadow_host = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "kat-modal[role='dialog'][aria-modal='true'][visible='true'] > span:nth-child(2) > kat-expander"))
        )
        INFO("Shadow host visible")

        shadow_root = driver.execute_script("return arguments[0].shadowRoot", shadow_host)
        if shadow_root:
            logging.info("Shadow root visible")
            kat_icon = shadow_root.find_element(By.CSS_SELECTOR, "div.wrapper > button > div.header__toggle > slot > kat-icon")
            kat_icon.click()
            logging.info("Kat icon clicked successfully.")
        else:
            WARNING("Shadow root not found.")
    except Exception as e:
        ERROR(f"An error occurred: {e}")

def extract_secret_key_and_expiration(driver):
    try:
        INFO("Attempting to copy the client secret...")

        # Attendre que le modal soit visible
        modal_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "kat-modal[role='dialog'][visible='true']"))
        )
        INFO("Modal visible")

        secret_div = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".clientSecretDiv kat-input"))
        )

        # Obtenir la valeur du secret client
        client_secret = secret_div.get_attribute("value")

        if client_secret:
            logging.info(f"Client secret successfully copied: {client_secret}")
        else:
            logging.error("Client secret is empty.")

        # Vérification de la date d'expiration
        logging.info("Attempting to verify the expiration date of the client secret...")

        # Attendre que la date d'expiration soit visible dans le modal
        expiration_span = WebDriverWait(modal_element, 10).until(
            EC.presence_of_element_located((By.XPATH, ".//span/div/i"))
        )
        logging.info("Expiration span found")

        # Obtenir la date d'expiration
        expiration_date_text = expiration_span.text
        expiration_date_str = expiration_date_text.split(":")[1].strip() if expiration_date_text else None

        if expiration_date_str:
            # Extraire uniquement la date au format YYYY-MM-DD
            expiration_date_formatted = expiration_date_str.split('T')[0]
            logging.info(f"Expiration date successfully extracted: {expiration_date_formatted}")
        else:
            logging.error("Expiration date is empty.")
            expiration_date_formatted = None

        return client_secret, expiration_date_formatted  # Retourner les deux valeurs

    except Exception as e:
        logging.error(f"Error extracting client secret and expiration: {str(e)}")
        return None, None  # Retourner None si une erreur se produit

def renew_secret_and_extract(driver, account_name, report_data):
    """Renew the secret and extract the new secret and expiration date."""
    try:
        # Attendez que le modal de renouvellement soit visible
        modal_element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "kat-modal[role='dialog'][aria-modal='true'][visible='true']"))
        )
        logging.info("Renewal modal is visible")

        # Cliquez sur le bouton de renouvellement
        renew_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "kat-button.footerLeftButton[variant='primary']"))
        )
        renew_button.click()
        logging.info("Renew secret button clicked")

        # Attendez que le bouton de confirmation soit visible
        confirm_button = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-button.footerConfirmationButton"))
        )
        confirm_button.click()
        logging.info("Confirmation button clicked")

        # Cliquez sur le bouton 'Terminé' sans utiliser l'attribut 'label'
        done_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "kat-button.footerRightButton"))
        )
        done_button.click()
        logging.info("Done button clicked.")

        # Cliquez à nouveau sur le bouton 'View' pour afficher le modal
        click_view_button(driver, account_name, report_data)

        # Attendez que le modal soit visible
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "kat-modal")))

        # Cliquez sur la flèche pour afficher la clé secrète
        click_arrow(driver)

        # Extraire le nouveau client secret et la nouvelle date d'expiration
        new_client_secret, new_expiration_date = extract_secret_key_and_expiration(driver)

        # Vérifiez si la nouvelle date d'expiration est valide
        if new_expiration_date:
            # Compléter la chaîne de date si nécessaire
            if len(new_expiration_date) < 10:  # "YYYY-MM-DD"
                new_expiration_date += 'T00:00:00'  # Ajouter le temps par défaut

            try:
                # Convertir la date d'expiration au format datetime
                new_expiration_datetime = datetime.fromisoformat(new_expiration_date.split('T')[0])  # Retirer l'heure
                days_until_new_expiration = (new_expiration_datetime - datetime.now()).days
            except ValueError as ve:
                logging.error(f"Error parsing new expiration date: {ve}")
                days_until_new_expiration = None
        else:
            days_until_new_expiration = None

        return new_client_secret, new_expiration_date

    except Exception as e:
        logging.error(f"Error during secret renewal: {e}")
        return None, None

def create_report(report_data):
    """Create an Excel report and return the filename."""
    current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_filename = REPORT_PATH_TEMPLATE.format(current_time)
    logging.info(f"Creating report at {report_filename}")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Secret Key Report"

    headers = [
        "Client Name",
        "Client ID",
        "Current Secret Key",
        "Current Expiration Date",
        "Days until Current Expiration",
        "New Secret Key",
        "New Expiration Date",
        "Days until New Expiration",
        "Renewal Status"
    ]
    sheet.append(headers)

    # Écrire les données dans le rapport
    for row in report_data:
        sheet.append(row)

        # Vérifiez la date d'expiration de la nouvelle clé secrète
        new_expiration_date = row[5]  # Nouvelle date d'expiration
        if new_expiration_date and new_expiration_date != "N/A":
            try:
                expiration_date = datetime.fromisoformat(new_expiration_date)
                days_until_expiration = (expiration_date - datetime.now()).days

                logging.info(f"Checking expiration for {row[0]}: {new_expiration_date}, Days until expiration: {days_until_expiration}")

                if days_until_expiration <= 30:
                    fill_color = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Rouge
                else:
                    fill_color = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Vert

                # Appliquer la couleur à la cellule de la date d'expiration
                sheet.cell(row=sheet.max_row, column=6).fill = fill_color  # Nouvelle date d'expiration
                # Optionnel : Appliquer la couleur à la cellule des jours jusqu'à l'expiration
                sheet.cell(row=sheet.max_row, column=7).fill = fill_color  # Jours jusqu'à la nouvelle expiration

            except Exception as e:
                logging.error(f"Error processing expiration date for {row[0]}: {e}")

    workbook.save(report_filename)

    return report_filename  # Return the filename to be used later

def main():
    # Lire les comptes depuis le fichier Excel
    accounts = read_accounts_from_excel(EXCEL_FILE_PATH)

    if not accounts:
        logging.error("Aucun compte trouvé dans le fichier Excel.")
        return

    # Initialiser le WebDriver
    driver = setup_driver()

    try:
        # Se connecter à Amazon
        login_to_amazon(driver)

        report_data = []  # Liste pour stocker les données du rapport

        for account in accounts:
            account_name = account['AccountName']
            target_client_id = account['ClientID']

            logging.info(f"Processing account: {account_name} (Client ID: {target_client_id})")

            # Sélectionner le compte du client
            select_client_account(driver, account_name)

            # Vérifier s'il y a une alerte sur le profil développeur
            if check_developer_profile_alert(driver):
                logging.warning("Developer profile alert detected. Handling the alert.")
                # Ici, vous pouvez ajouter le code pour gérer l'alerte si nécessaire.

            # Trouver l'application par son Client ID
            if find_application_by_client_id(driver, target_client_id):
                # Extraire la clé secrète et la date d'expiration
                client_secret, expiration_date = extract_secret_key_and_expiration(driver)

                if client_secret is None or expiration_date is None:
                    logging.warning(f"Skipping account {account_name}, missing key or expiration date.")
                    continue

                # Calculer les jours restants avant expiration
                days_until_expiration = (datetime.fromisoformat(expiration_date) - datetime.now()).days if expiration_date else None
                logging.info(f"Days until expiration: {days_until_expiration} for {account_name}")

                # Si la date d'expiration est inférieure à 30 jours, renouveler la clé
                if days_until_expiration and days_until_expiration < 30:
                    logging.info(f"Renewing secret for account: {account_name}")
                    new_client_secret, new_expiration_date = renew_secret_and_extract(driver, account_name, report_data)
                    report_data.append([account_name,target_client_id, client_secret, expiration_date, days_until_expiration, new_client_secret, new_expiration_date, days_until_expiration, "Renewed successfully"])
                else:
                    report_data.append([account_name, target_client_id, client_secret, expiration_date, days_until_expiration, "N/A", "N/A", "N/A", "No renewal needed"])
            else:
                logging.warning(f"No application found for ClientID: {target_client_id} in account {account_name}")

        # Générer le rapport avec les données collectées
        if report_data:
            report_filename = create_report(report_data)
            logging.info(f"Report generated: {report_filename}")
        else:
            logging.warning("No data to report.")

    except Exception as e:
        logging.error(f"An error occurred during the process: {e}")
    finally:
        # Fermer le driver
        driver.quit()

if __name__ == "__main__":
    main()